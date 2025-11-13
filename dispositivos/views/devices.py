# dispositivos/views/devices.py
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST

from dispositivos.models import Device
from dispositivos.forms import DeviceForm
from monitoreo.decorators import permission_or_redirect


# ─────────────────────────────────────────────────────────────────────────────
# LISTA AVANZADA: buscador, orden y paginación con persistencia
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def device_list_advanced(request):
    q = (request.GET.get("q") or "").strip()
    sort = request.GET.get("sort", "name")

    # Persistir tamaño de página en sesión
    if "page_size" in request.GET:
        try:
            request.session["device_page_size"] = int(request.GET.get("page_size") or 10)
        except ValueError:
            request.session["device_page_size"] = 10
    page_size = int(request.session.get("device_page_size", 10))

    qs = (
        Device.objects.select_related("organization", "zone", "product", "product__category")
        .order_by(sort)
    )

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(organization__name__icontains=q) |
            Q(zone__name__icontains=q) |
            Q(product__name__icontains=q) |
            Q(product__category__name__icontains=q)
        )

    paginator = Paginator(qs, page_size)
    page_number = request.GET.get("page")
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    params = request.GET.copy()
    params.pop("page", None)
    querystring = params.urlencode()

    ctx = {
        "page_obj": page_obj,
        "q": q,
        "sort": sort,
        "page_size": page_size,
        "querystring": querystring,
        "total": qs.count(),
        "page_sizes": [5, 10, 25, 50],
    }
    return render(request, "devices/devices_list_paginador.html", ctx)


# ─────────────────────────────────────────────────────────────────────────────
# CREATE - con validaciones mejoradas
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@permission_or_redirect("dispositivos.add_device", "device_list_advanced", "No tienes permiso para crear dispositivos.")
def device_create(request):
    if request.method == "POST":
        form = DeviceForm(request.POST, request.FILES)
        if form.is_valid():
            # Validación adicional: verificar duplicados (nombre + zona)
            nombre = form.cleaned_data.get('name')
            zona = form.cleaned_data.get('zone')
            
            if Device.objects.filter(name__iexact=nombre, zone=zona).exists():
                messages.error(request, f'❌ Ya existe un dispositivo llamado "{nombre}" en la zona "{zona.name}".')
                return render(request, "devices/device_form.html", {"form": form, "accion": "Crear"})
            
            # Validación: longitud mínima
            if len(nombre) < 3:
                messages.error(request, '❌ El nombre debe tener al menos 3 caracteres.')
                return render(request, "devices/device_form.html", {"form": form, "accion": "Crear"})
            
            # Validación: potencia máxima positiva (si existe)
            max_power = form.cleaned_data.get('max_power_w')
            if max_power is not None and max_power <= 0:
                messages.error(request, '❌ La potencia máxima debe ser un valor positivo.')
                return render(request, "devices/device_form.html", {"form": form, "accion": "Crear"})
            
            form.save()
            messages.success(request, f'✅ Dispositivo "{nombre}" creado exitosamente.')
            return redirect("device_list_advanced")
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f'❌ {error}')
                    else:
                        field_label = form.fields[field].label or field
                        messages.error(request, f'❌ {field_label}: {error}')
    else:
        form = DeviceForm()
    
    return render(request, "devices/device_form.html", {"form": form, "accion": "Crear"})


# ─────────────────────────────────────────────────────────────────────────────
# UPDATE - con validaciones mejoradas
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@permission_or_redirect("dispositivos.change_device", "device_list_advanced", "No tienes permiso para editar dispositivos.")
def device_edit(request, pk):
    device = get_object_or_404(Device, pk=pk)
    
    if request.method == "POST":
        form = DeviceForm(request.POST, request.FILES, instance=device)
        if form.is_valid():
            # Validación adicional: verificar duplicados (excluyendo el dispositivo actual)
            nombre = form.cleaned_data.get('name')
            zona = form.cleaned_data.get('zone')
            
            duplicado = Device.objects.filter(
                name__iexact=nombre, 
                zone=zona
            ).exclude(pk=pk).exists()
            
            if duplicado:
                messages.error(request, f'❌ Ya existe otro dispositivo llamado "{nombre}" en la zona "{zona.name}".')
                return render(request, "devices/device_form.html", {"form": form, "accion": "Editar"})
            
            # Validación: longitud mínima
            if len(nombre) < 3:
                messages.error(request, '❌ El nombre debe tener al menos 3 caracteres.')
                return render(request, "devices/device_form.html", {"form": form, "accion": "Editar"})
            
            # Validación: potencia máxima positiva (si existe)
            max_power = form.cleaned_data.get('max_power_w')
            if max_power is not None and max_power <= 0:
                messages.error(request, '❌ La potencia máxima debe ser un valor positivo.')
                return render(request, "devices/device_form.html", {"form": form, "accion": "Editar"})
            
            form.save()
            messages.success(request, f'✅ Dispositivo "{nombre}" actualizado correctamente.')
            return redirect("device_list_advanced")
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f'❌ {error}')
                    else:
                        field_label = form.fields[field].label or field
                        messages.error(request, f'❌ {field_label}: {error}')
    else:
        form = DeviceForm(instance=device)
    
    return render(request, "devices/device_form.html", {"form": form, "accion": "Editar"})


# ─────────────────────────────────────────────────────────────────────────────
# DELETE (AJAX + SweetAlert2)
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@require_POST
def device_delete_ajax(request, pk):
    if not request.user.has_perm("dispositivos.delete_device"):
        return JsonResponse({"ok": False, "message": "❌ No tienes permiso para eliminar dispositivos."}, status=403)

    if request.headers.get("x-requested-with") != "XMLHttpRequest":
        return HttpResponseBadRequest("Solo AJAX")

    device = get_object_or_404(Device, pk=pk)
    nombre = getattr(device, "name", f"ID {device.pk}")
    
    # Verificar si el dispositivo tiene mediciones asociadas
    if hasattr(device, 'measurement_set') and device.measurement_set.exists():
        count = device.measurement_set.count()
        return JsonResponse({
            "ok": False, 
            "message": f"❌ No se puede eliminar el dispositivo '{nombre}' porque tiene {count} medición(es) asociada(s)."
        }, status=400)
    
    device.delete()
    return JsonResponse({"ok": True, "message": f"✅ Dispositivo '{nombre}' eliminado exitosamente."})


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL (xlsx) con filtros/orden actuales
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def device_export_xlsx(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    q = (request.GET.get("q") or "").strip()
    sort = request.GET.get("sort", "name")

    qs = (
        Device.objects.select_related("organization", "zone", "product", "product__category")
        .order_by(sort)
    )
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(organization__name__icontains=q) |
            Q(zone__name__icontains=q) |
            Q(product__name__icontains=q) |
            Q(product__category__name__icontains=q)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dispositivos"

    # Encabezados con estilo
    headers = ["ID", "Nombre", "Número de Serie", "Producto", "Categoría", "Zona", "Organización", "Potencia máx. (W)"]
    ws.append(headers)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for d in qs:
        ws.append([
            d.id,
            d.name,
            d.serial_number or "",
            d.product.name if d.product_id else "",
            d.product.category.name if d.product_id and d.product.category_id else "",
            d.zone.name if d.zone_id else "",
            d.organization.name if d.organization_id else "",
            d.max_power_w or 0,
        ])

    # Ancho de columnas automático
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(12, length + 2), 50)

    # Respuesta HTTP
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="dispositivos_ecoenergy.xlsx"'
    wb.save(resp)
    return resp