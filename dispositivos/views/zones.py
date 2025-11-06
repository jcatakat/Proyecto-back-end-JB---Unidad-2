# dispositivos/views/zones.py
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.http import (
    JsonResponse,
    HttpResponseBadRequest,
    HttpResponse,
)
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST

from dispositivos.models import Zone
from dispositivos.forms import ZoneForm
from monitoreo.decorators import permission_or_redirect


# ─────────────────────────────────────────────────────────────────────────────
# LISTA: buscador (q), orden (sort), y paginación con page_size recordado
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def zona_list(request):
    q = (request.GET.get("q") or "").strip()
    sort = request.GET.get("sort", "name")  # "name" o "organization__name"

    # Persistir tamaño de página en sesión (default 10)
    if "page_size" in request.GET:
        try:
            request.session["zona_page_size"] = int(request.GET.get("page_size") or 10)
        except ValueError:
            request.session["zona_page_size"] = 10
    page_size = int(request.session.get("zona_page_size", 10))

    qs = Zone.objects.select_related("organization").order_by(sort)
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(organization__name__icontains=q)
        )

    paginator = Paginator(qs, page_size)
    page_number = request.GET.get("page")

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Conservar filtros en enlaces del paginador (q, sort, page_size)
    params = request.GET.copy()
    params.pop("page", None)
    querystring = params.urlencode()

    ctx = {
        "page_obj": page_obj,
        "q": q,
        "sort": sort,
        "page_size": page_size,
        "querystring": querystring,
        "total": qs.count(),
        "page_sizes": [5, 10, 25, 50], 
    }
    # Usa el template con paginación/buscador
    return render(request, "zona/zonas_list_paginador.html", ctx)


# ─────────────────────────────────────────────────────────────────────────────
# CREATE
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@permission_or_redirect("dispositivos.add_zone", "zona_list", "No tienes permiso para crear zonas.")
def zona_create(request):
    if request.method == "POST":
        form = ZoneForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Zona creada exitosamente")
            return redirect("zona_list")
        messages.error(request, "Corrige los errores")
    else:
        form = ZoneForm()
    return render(request, "zona/zona_form.html", {"form": form, "accion": "Crear"})


# ─────────────────────────────────────────────────────────────────────────────
# UPDATE
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@permission_or_redirect("dispositivos.change_zone", "zona_list", "No tienes permiso para editar zonas.")
def zona_edit(request, pk):
    zona = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        form = ZoneForm(request.POST, instance=zona)
        if form.is_valid():
            form.save()
            messages.success(request, "Zona actualizada correctamente")
            return redirect("zona_list")
        messages.error(request, "Corrige los errores")
    else:
        form = ZoneForm(instance=zona)
    return render(request, "zona/zona_form.html", {"form": form, "accion": "Editar"})


# ─────────────────────────────────────────────────────────────────────────────
# DELETE (AJAX + SweetAlert2)
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@require_POST
def zona_delete_ajax(request, pk):
    if not request.user.has_perm("dispositivos.delete_zone"):
        return JsonResponse({"ok": False, "message": "No tienes permiso para eliminar."}, status=403)

    if request.headers.get("x-requested-with") != "XMLHttpRequest":
        return HttpResponseBadRequest("Solo AJAX")

    zona = get_object_or_404(Zone, pk=pk)
    nombre = zona.name
    zona.delete()
    return JsonResponse({"ok": True, "message": f"Zona '{nombre}' eliminada"})


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL (xlsx) con filtros/orden actuales
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def zona_export_xlsx(request):
    import openpyxl
    from openpyxl.utils import get_column_letter

    q = (request.GET.get("q") or "").strip()
    sort = request.GET.get("sort", "name")

    qs = Zone.objects.select_related("organization").order_by(sort)
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(organization__name__icontains=q)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zonas"

    headers = ["ID", "Nombre", "Organización"]
    ws.append(headers)

    for z in qs:
        ws.append([z.id, z.name, getattr(z.organization, "name", "")])

    # Ancho de columnas simple
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(10, length + 2), 40)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="zonas.xlsx"'
    wb.save(resp)
    return resp
